import requests
import pandas as pd
from datetime import datetime
from win32com import client
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.label import DataLabelList
import fitz

dir_path = os.path.dirname(os.path.realpath('__file__'))

def dados(x):
    r = requests.get(x)
    return r.json()

def imprimir(n, nome):
    excel = client.Dispatch("Excel.Application")
    sheets = excel.Workbooks.Open(f'{dir_path}\\{nome}.xlsx')
    work_sheets = sheets.Worksheets[n]
    work_sheets.ExportAsFixedFormat(0, f'{dir_path}\\{nome}.pdf')
    sheets.Close(True)
    doc = fitz.open(f'{dir_path}\\{nome}.pdf')
    page = doc.load_page(0)
    pix = page.get_pixmap()
    pix.save(f'{dir_path}\\{nome}.png')

class Tabela():
    
    def __init__(self, df=pd.DataFrame()):
        self._df = df
        
    @property
    def df(self):
        return self._df

    def criar(self, r, lista=[]):
        df = self._df
        n = 0
        if lista == []:
            df = pd.DataFrame(r)
        else:
            for m in lista:
                dfr = pd.DataFrame(r[m], index=[n])
                df = pd.concat([df, dfr])
                n += 1
        self._df = df

    def renomear_coluna(self, dicionario):
        self._df = self._df.rename(columns=dicionario)
    
    def coluna_num(self, lista):
        df = self._df
        for num in lista:
            df[num] = pd.to_numeric(self._df[num])
        self._df = df

    def coluna_data(self, lista):
        df = self._df
        for coluna in lista:
            df[coluna] = pd.to_datetime(self._df[coluna])
        self._df = df
    
    def excluir_coluna(self, lista):
        df = self._df
        for coluna in lista: 
            df = self._df.drop(columns=coluna)
        self._df = df
        
    def duplicar_info(self, lista):
        df = self._df
        for coluna in lista: 
            df[coluna] = self._df[coluna].fillna(method='ffill')
        self._df = df
        
    def converter_data(self, lista):
        df = self._df
        for coluna in lista: 
            df[coluna] = self._df[coluna].apply(lambda x: datetime.fromtimestamp(int(x)))
        self._df = df

class Excel():
    
    def __init__(self, nome):
        self._nome = nome
        
    def pasta(self):
        wb = Workbook()
        return wb
        
    def planilha(self, wb, nome, index=None):
        return wb.create_sheet(nome, index)
    
    def excluir_planilha(self, wb, nome):
        std = wb.get_sheet_by_name(nome)
        wb.remove_sheet(std)
    
    def tabela(self, df, ws):
        # Inserir tabela
        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)
        # Formatar tabela
        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'
        # Remover celulas em branco
        def remove(ws, row): 
            for cell in row: 
                if cell.value != None: 
                    return
            ws.delete_rows(row[0].row, 1) 
        for row in ws: 
            remove(ws,row)
        ws.delete_cols(1)
        
    def alinhar(self, ws, n=2):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
        for col, value in dims.items():
            ws.column_dimensions[col].width = (value + 2)
    
    def form_data(self, ws, lista):
        for s in lista:
            for cell in ws[s]:
                cell.number_format = "DD/MM/YYYY hh:mm:ss"
                
    def form_porcentagem(self, ws, lista):
        for s in lista:
            for cell in ws[s]:
                cell.number_format = "0.00%"
    
    def form_centralizar(self, ws, lista):
        for s in lista:
            for cell in ws[s]:
                cell.alignment = Alignment(horizontal='center')
                
    def salvar(self, wb):
        wb.save(f"{self._nome}.xlsx")
        
    def ocultar_planilha(self, planilha):
        planilha.sheet_state = 'hidden'
        
    def estilo(self, planilha, celula, tamanho, cor, negrito, alinhamento, borda=True):
        planilha[celula].font = Font(size=tamanho, b=negrito)
        if alinhamento == True:
            planilha[celula].alignment = Alignment(horizontal="center", vertical="center")
        planilha[celula].fill = PatternFill("solid", fgColor=cor)
        if borda == True:
            thin = Side(border_style="thick", color="000000")
            planilha[celula].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
    def mesclar(self, planilha, intervalo):
        planilha.merge_cells(intervalo)
        
    def escrever(self, planilha, celula, texto):
        planilha[celula].value = texto
        
    def inserir_linha(self, planilha, numero, quantidade):
        for n in range(0, quantidade):
            planilha.insert_rows(numero)
                
    def inserir_coluna(self, planilha, numero, quantidade):
        for n in range(0, quantidade):
            planilha.insert_cols(numero)
        
    def grafico_linha(self, grafico, dados, df, titulo, x, y, c, n="#,##0.00"):
        df = df[[x, y]].sort_values(by=[x],ignore_index=True, ascending=True)
        for r in dataframe_to_rows(df, index=True, header=True):
            dados.append(r)
        for cell in dados['A'] + dados[1]:
            cell.style = 'Pandas'
        # Remover celulas em branco
        def remove(dados, row): 
            for cell in row: 
                if cell.value != None: 
                    return
            dados.delete_rows(row[0].row, 1) 
        for row in dados: 
            remove(dados,row)
        dados.delete_cols(1)
        
        dims = {}
        for row in dados.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
        for col, value in dims.items():
            dados.column_dimensions[col].width = (value + 2)
            
        for cell in dados["A"]:
            cell.number_format = "DD-MM"
            
        for cell in dados["B"]:
            cell.number_format = n
            
        for s in ["A", "B"]:
            for cell in dados[s]:
                cell.alignment = Alignment(horizontal='center')

        c1 = LineChart()
        c1.title = titulo
        c1.style = 3
        c1.y_axis.title = y
        c1.x_axis.title = x
        c1.dataLabels = DataLabelList() 
        c1.dataLabels.showVal = True
        
        data = Reference(dados, min_col=2, min_row=1, max_col=2, max_row=(len(df.index)+1))
        c1.add_data(data, titles_from_data=True)
        dates = Reference(dados, min_col=1, min_row=2, max_col=1, max_row=(len(df.index)+1))
        c1.set_categories(dates)
        
        grafico.add_chart(c1, c)
        
    def impressora(self, planilha):
        planilha.sheet_properties.pageSetUpPr.fitToPage = True
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(planilha, paper_size = 1, orientation='landscape')

    